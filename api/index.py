from flask import Flask, request, send_file
import openpyxl
import io
import os
import json
import datetime

app = Flask(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

@app.route('/api/process', methods=['POST'])
def process_excel():
    try:
        if 'file' not in request.files:
            return {"error": "No file uploaded"}, 400
        
        uploaded_file = request.files['file']
        # Web画面からの入力データ
        table_data = json.loads(request.form.get('tableData'))

        # テンプレート読み込み
        with open(TEMPLATE_PATH, 'rb') as f:
            template_buffer = io.BytesIO(f.read())
        wb_template = openpyxl.load_workbook(template_buffer)
        
        # =========================================================
        # 処理A: データ解析フェーズ (Pythonで値を全部計算してしまう)
        # =========================================================
        wb_uploaded = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws_src = wb_uploaded.worksheets[0] # 1枚目のシート
        
        # 全行をリストとして取得（行番号でアクセスするため）
        # min_row=1 から最後まで取得
        all_rows = list(ws_src.iter_rows(values_only=True))

        unique_names = []
        arrival_data = {}   # 登園（上の段）
        departure_data = {} # 降園（下の段）

        # データ開始行（以前の画像から58行目付近と推測されますが、安全のため全体走査または固定）
        # ここではループで見つけます
        START_ROW_INDEX = 57 # 58行目 = index 57

        for i, row in enumerate(all_rows):
            # 開始行より前はスキップ
            if i < START_ROW_INDEX:
                continue
            
            # A列(index 0)が名前
            name = row[0]
            
            # 名前が無効ならスキップ
            if not name or name == "お子さま名" or str(name) == "0":
                continue

            # 名前リスト作成
            if name not in unique_names:
                unique_names.append(name)
                arrival_data[name] = {}
                departure_data[name] = {}

            # -----------------------------------------------------
            # 登園時間の取得（名前と同じ行：F列～）
            # -----------------------------------------------------
            # F列は index 5。ここから31日分(index 35まで)を見る
            for day in range(31):
                col_idx = 5 + day
                if col_idx < len(row):
                    val = row[col_idx]
                    # 0や空でなければ採用
                    if val and val != 0:
                        arrival_data[name][day + 1] = val

            # -----------------------------------------------------
            # 降園時間の取得（名前の【1つ下の行】：F列～）
            # -----------------------------------------------------
            # 次の行が存在するか確認
            if i + 1 < len(all_rows):
                next_row = all_rows[i + 1]
                for day in range(31):
                    col_idx = 5 + day
                    if col_idx < len(next_row):
                        val = next_row[col_idx]
                        # 0や空でなければ採用
                        if val and val != 0:
                            departure_data[name][day + 1] = val

        # =========================================================
        # 処理B: 書き込みフェーズ
        # =========================================================

        # 1. 「貼り付け用」シート（1枚目）へ、アップロードデータをそのままコピー
        ws_paste = wb_template['貼り付け用']
        for i, row in enumerate(all_rows, start=1):
            for j, value in enumerate(row, start=1):
                ws_paste.cell(row=i, column=j, value=value)

        # 2. 「子どもマスタ」シート（2枚目）へ、★Web入力データ★ を貼る
        # （ここを元に戻しました）
        if '子どもマスタ' in wb_template.sheetnames:
            ws_child = wb_template['子どもマスタ']
            for row_idx, row_data in enumerate(table_data):
                for col_idx, value in enumerate(row_data):
                    # A2から貼り付け
                    ws_child.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        # 3. 「まとめ（登園）」シート（3枚目）へ、★Python計算値(登園)★ を貼る
        if 'まとめ（登園）' in wb_template.sheetnames:
            ws_arrival = wb_template['まとめ（登園）']
            # B3からスタート
            BASE_ROW = 3
            
            for idx, name in enumerate(unique_names):
                current_row = BASE_ROW + idx
                # B列: 名前
                ws_arrival.cell(row=current_row, column=2, value=name)
                
                # F列～: 時間
                if name in arrival_data:
                    days = arrival_data[name]
                    for day, time_val in days.items():
                        # 1日=F列(6列目) なので、 column = 5 + day
                        ws_arrival.cell(row=current_row, column=5 + day, value=time_val)

        # 4. 「まとめ（降園）」シート（4枚目）へ、★Python計算値(降園)★ を貼る
        if 'まとめ（降園）' in wb_template.sheetnames:
            ws_departure = wb_template['まとめ（降園）']
            # B3からスタート
            BASE_ROW = 3
            
            for idx, name in enumerate(unique_names):
                current_row = BASE_ROW + idx
                # B列: 名前
                ws_departure.cell(row=current_row, column=2, value=name)
                
                # F列～: 時間
                if name in departure_data:
                    days = departure_data[name]
                    for day, time_val in days.items():
                        ws_departure.cell(row=current_row, column=5 + day, value=time_val)

        # 保存処理
        output_stream = io.BytesIO()
        wb_template.save(output_stream)
        output_stream.seek(0)

        today_str = datetime.datetime.now().strftime('%Y%m%d')
        download_name = f"complete_{today_str}.xlsx"

        return send_file(
            output_stream,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return {"error": str(e)}, 500
